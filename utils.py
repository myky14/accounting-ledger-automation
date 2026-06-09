from datetime import date, datetime
from io import BytesIO
import math
from pathlib import Path
import re
import warnings

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


APPROVED_STATUS = "Approved"
APPROVAL_STATUS_VALUES = [
    APPROVED_STATUS,
    "Rejected",
    "Needs Review",
    "Employee Not Found",
    "Vendor Not Found",
    "Ignore",
]

MATCH_STATUS_VALUES = [
    "AUTO MATCH",
    "LOW CONFIDENCE",
    "AMBIGUOUS",
    "NO MATCH",
    "CONTEXT MATCH",
    "LOAN OUT MATCH",
    "DIRECT VENDOR MATCH",
]


def safe_value(value):
    """Return a blank string for missing spreadsheet values."""
    if pd.isna(value):
        return ""
    return value


def safe_excel_text(value):
    # Tax ID/SIN/GST/Zip Code phải xử lý như text để không mất số 0 đầu.
    """Return a text-safe Excel value for IDs, zip codes, SIN, GST, and addresses."""
    # Tax ID, SIN, Zip Code nhìn giống số nhưng phải xử lý như text để không mất số 0 đầu và tránh lỗi dtype.
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def is_blank(value):
    return pd.isna(value) or str(value).strip() == ""


def normalize_name(name):
    """Normalize names so small formatting differences do not block matching."""
    if pd.isna(name):
        return ""
    name = str(name).upper()
    name = re.sub(r"[^A-Z0-9 ]", "", name)
    return name.replace(" ", "")


def ensure_input_file(path_value, label):
    path = Path(path_value)
    if not path.exists():
        raise FileNotFoundError(f"{label} does not exist: {path}")
    return path


def ensure_parent_dir(path_value):
    Path(path_value).parent.mkdir(parents=True, exist_ok=True)


def read_workbook(path_value, sheet_name=0, dtype=None):
    path = ensure_input_file(path_value, "Input file")
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=dtype)
    return pd.read_excel(path, sheet_name=sheet_name, dtype=dtype)


def sanitize_sheet_name(name, existing_names=None):
    """Return an Excel-safe worksheet name.

    Excel chỉ cho sheet name tối đa 31 ký tự và không cho các ký tự
    colon, backslash, slash, question mark, star, bracket. Nếu vượt giới hạn,
    Microsoft Excel có thể repair workbook.
    """
    existing_names = existing_names or set()
    cleaned_name = str(name or "Sheet").strip()
    cleaned_name = re.sub(r"[:\\/\?\*\[\]]", " ", cleaned_name)
    cleaned_name = re.sub(r"\s+", " ", cleaned_name).strip() or "Sheet"
    cleaned_name = cleaned_name[:31]

    candidate = cleaned_name
    counter = 1
    while candidate in existing_names:
        suffix = f" {counter}"
        candidate = f"{cleaned_name[:31 - len(suffix)]}{suffix}"
        counter += 1
    existing_names.add(candidate)
    return candidate


def clean_excel_value(value):
    """Clean one value before writing it to Excel.

    Excel không chấp nhận một số control characters trong cell text. NaN/inf
    cũng nên đổi thành blank để workbook mở ổn định trong Microsoft Excel.
    """
    if value is None or pd.isna(value):
        return ""

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return value

    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", value)

    return value


def clean_dataframe_for_excel(df):
    """Return a copy of df with Excel-unsafe values removed."""
    if df is None:
        return pd.DataFrame()
    cleaned_df = df.copy()
    cleaned_df = cleaned_df.replace([float("inf"), float("-inf")], "")
    cleaned_df = cleaned_df.fillna("")
    return cleaned_df.map(clean_excel_value)


def write_workbook_safely(output_path_or_buffer, sheets_dict):
    """Write an xlsx workbook with safe sheet names and safe cell values.

    Nếu output là BytesIO, writer phải đóng xong và buffer phải seek(0) trước
    khi đưa cho Streamlit download_button. Nếu không, file download có thể bị
    thiếu bytes hoặc bị Excel repair.
    """
    if not isinstance(output_path_or_buffer, BytesIO):
        ensure_parent_dir(output_path_or_buffer)

    existing_sheet_names = set()
    with pd.ExcelWriter(output_path_or_buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            safe_sheet_name = sanitize_sheet_name(sheet_name, existing_sheet_names)
            safe_df = clean_dataframe_for_excel(df)
            safe_df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

    if isinstance(output_path_or_buffer, BytesIO):
        output_path_or_buffer.seek(0)

    validate_workbook(output_path_or_buffer)

    if isinstance(output_path_or_buffer, BytesIO):
        output_path_or_buffer.seek(0)


def apply_review_workbook_formatting(output_path, review_sheet_names=None):
    """Add review-friendly dropdowns, comments, and row highlighting."""
    if isinstance(output_path, BytesIO):
        return

    workbook = load_workbook(output_path)
    requested_names = set(review_sheet_names or [])

    for worksheet in workbook.worksheets:
        original_title = worksheet.title
        if requested_names and not any(
            original_title == name or original_title.startswith(name[:31])
            for name in requested_names
        ):
            continue
        _format_review_sheet(worksheet)

    workbook.save(output_path)
    workbook.close()
    validate_workbook(output_path)


def _format_review_sheet(worksheet):
    if worksheet.max_row < 1:
        return

    headers = {
        str(cell.value).strip(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    approval_column = headers.get("Approval Status")
    match_column = headers.get("Match Status")
    notes_column = headers.get("Reviewer Notes")

    # Match Status là ý kiến của algorithm.
    # Approval Status mới là authority cuối cùng.
    if match_column:
        worksheet.cell(row=1, column=match_column).comment = Comment(
            "System-generated suggestion status. This is not final approval.",
            "Ledger Automation",
        )
    if approval_column:
        worksheet.cell(row=1, column=approval_column).comment = Comment(
            "Manual reviewer decision. Final enrichment only runs when this is Approved.",
            "Ledger Automation",
        )
        _add_approval_dropdown(worksheet, approval_column)
        _add_status_row_highlighting(worksheet, approval_column, match_column)
    if notes_column:
        worksheet.cell(row=1, column=notes_column).comment = Comment(
            "Use Reviewer Notes to explain why a match was approved or rejected. "
            "This helps audit, debugging, and future alias-memory improvements.",
            "Ledger Automation",
        )

    for column_cells in worksheet.columns:
        header = str(column_cells[0].value or "")
        max_length = max(
            len(str(cell.value or "")) for cell in column_cells[: min(len(column_cells), 100)]
        )
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_length + 2, len(header) + 2, 12),
            42,
        )


def _add_approval_dropdown(worksheet, approval_column):
    values_formula = '"' + ",".join(APPROVAL_STATUS_VALUES) + '"'
    validation = DataValidation(
        type="list",
        formula1=values_formula,
        allow_blank=True,
    )
    validation.error = "Choose a valid Approval Status from the dropdown."
    validation.errorTitle = "Invalid Approval Status"
    validation.prompt = "Only Approved rows can enrich Tax ID/address."
    validation.promptTitle = "Manual approval required"
    worksheet.add_data_validation(validation)

    approval_letter = get_column_letter(approval_column)
    last_row = max(worksheet.max_row, 1000)
    validation.add(f"{approval_letter}2:{approval_letter}{last_row}")


def _add_status_row_highlighting(worksheet, approval_column, match_column=None):
    last_row = max(worksheet.max_row, 2)
    last_column = get_column_letter(worksheet.max_column)
    approval_letter = get_column_letter(approval_column)
    range_ref = f"A2:{last_column}{last_row}"

    green_fill = PatternFill("solid", fgColor="D9EAD3")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    red_fill = PatternFill("solid", fgColor="F4CCCC")

    worksheet.conditional_formatting.add(
        range_ref,
        FormulaRule(
            formula=[f'${approval_letter}2="{APPROVED_STATUS}"'],
            fill=green_fill,
        ),
    )
    worksheet.conditional_formatting.add(
        range_ref,
        FormulaRule(
            formula=[f'${approval_letter}2="Needs Review"'],
            fill=yellow_fill,
        ),
    )

    red_conditions = [
        f'${approval_letter}2="Rejected"',
        f'${approval_letter}2="Employee Not Found"',
        f'${approval_letter}2="Vendor Not Found"',
    ]
    if match_column:
        match_letter = get_column_letter(match_column)
        red_conditions.append(f'${match_letter}2="AMBIGUOUS"')

    worksheet.conditional_formatting.add(
        range_ref,
        FormulaRule(formula=[f"OR({','.join(red_conditions)})"], fill=red_fill),
    )


def validate_workbook(path_or_buffer):
    """Reload workbook with openpyxl so corruption is caught before download."""
    try:
        if isinstance(path_or_buffer, BytesIO):
            current_position = path_or_buffer.tell()
            path_or_buffer.seek(0)
            workbook = load_workbook(path_or_buffer, read_only=True, data_only=False)
            workbook.close()
            path_or_buffer.seek(current_position)
        else:
            workbook = load_workbook(path_or_buffer, read_only=True, data_only=False)
            workbook.close()
    except Exception as exc:
        raise ValueError(f"Workbook validation failed: {exc}") from exc


def resolve_column(df, column_ref, label):
    """Resolve a configured column name or zero-based index to a real df column."""
    if column_ref is None or column_ref == "":
        return None

    if isinstance(column_ref, list):
        raise ValueError(
            f"{label} must be one column name, but got a list: {column_ref}. "
            "Use a list-aware config field and loop over each column instead."
        )

    if not isinstance(column_ref, (str, int)):
        raise ValueError(
            f"{label} must be a column name string or zero-based column index, "
            f"but got {type(column_ref).__name__}: {column_ref}"
        )

    if isinstance(column_ref, int):
        if column_ref < 0 or column_ref >= len(df.columns):
            raise ValueError(
                f"{label} index {column_ref} is outside available columns "
                f"0..{len(df.columns) - 1}"
            )
        return df.columns[column_ref]

    if column_ref not in df.columns:
        stripped_matches = [
            column
            for column in df.columns
            if str(column).strip() == str(column_ref).strip()
        ]
        if len(stripped_matches) == 1:
            return stripped_matches[0]
        if len(stripped_matches) > 1:
            raise ValueError(
                f"{label} column '{column_ref}' matched multiple columns after "
                f"trimming whitespace: {stripped_matches}"
            )
        available = ", ".join(str(col) for col in df.columns)
        raise ValueError(
            f"{label} column '{column_ref}' was not found. Available columns: {available}"
        )
    return column_ref


def validate_master_columns(df, required_columns, optional_columns=None, label="Master"):
    """Validate master-data headers with whitespace-tolerant matching."""
    required_columns = required_columns or []
    optional_columns = optional_columns or []
    available_stripped = {str(column).strip() for column in df.columns}
    missing_required = [
        column for column in required_columns if str(column).strip() not in available_stripped
    ]
    if missing_required:
        available = ", ".join(str(column) for column in df.columns)
        raise ValueError(
            f"{label} is missing required columns: "
            + ", ".join(missing_required)
            + f". Available columns: {available}"
        )

    missing_optional = [
        column for column in optional_columns if str(column).strip() not in available_stripped
    ]
    if missing_optional:
        warnings.warn(
            f"{label} is missing optional columns: {', '.join(missing_optional)}",
            stacklevel=2,
        )
    return missing_optional


def resolve_column_list(df, column_refs, label, allow_missing=False):
    """Resolve a config field that intentionally accepts multiple columns."""
    if column_refs is None or column_refs == "":
        return []
    if isinstance(column_refs, (str, int)):
        column_refs = [column_refs]
    if not isinstance(column_refs, list):
        raise ValueError(
            f"{label} must be a list of column names, but got "
            f"{type(column_refs).__name__}: {column_refs}"
        )

    resolved = []
    for index, column_ref in enumerate(column_refs):
        if allow_missing and (column_ref is None or column_ref == ""):
            continue
        resolved.append(resolve_column(df, column_ref, f"{label}[{index}]"))
    return resolved


def resolve_column_map(df, mapping, label, allow_missing=False):
    resolved = {}
    for key, column_ref in mapping.items():
        if allow_missing and (column_ref is None or column_ref == ""):
            resolved[key] = None
            continue
        resolved[key] = resolve_column(df, column_ref, f"{label}.{key}")
    return resolved


def get_cell(row, resolved_columns, key):
    column = resolved_columns.get(key)
    if column is None:
        return ""
    return safe_value(row[column])


def row_snapshot(row):
    """Create a compact text version of a raw row for audit/review sheets."""
    pieces = []
    for column, value in row.items():
        if not is_blank(value):
            pieces.append(f"{column}={safe_value(value)}")
    return " | ".join(pieces)


def is_transaction_date(value):
    """Return True only for values that look like real transaction dates.

    This avoids the dangerous pattern where values like "0401" or "443131.68"
    are accidentally parsed as dates by pandas.
    """
    if is_blank(value):
        return False

    if isinstance(value, (pd.Timestamp, datetime, date)):
        return not pd.isna(value)

    if isinstance(value, (int, float)):
        return False

    value_text = str(value).strip()
    lower_text = value_text.lower()
    if "total" in lower_text or "subtotal" in lower_text or "expense" in lower_text:
        return False

    date_like_patterns = [
        r"^\d{4}-\d{1,2}-\d{1,2}$",
        r"^\d{4}/\d{1,2}/\d{1,2}$",
        r"^\d{1,2}/\d{1,2}/\d{2,4}$",
        r"^\d{1,2}-\d{1,2}-\d{2,4}$",
    ]
    if not any(re.match(pattern, value_text) for pattern in date_like_patterns):
        return False

    parsed = pd.to_datetime(value_text, errors="coerce")
    return not pd.isna(parsed)


def amount_to_number(value):
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").replace("$", "").strip()
    parsed = pd.to_numeric(cleaned, errors="coerce")
    if pd.isna(parsed):
        return None
    return float(parsed)


def contains_any_keyword(row, keywords):
    text = " ".join(str(value).lower() for value in row.values if not is_blank(value))
    return any(str(keyword).lower() in text for keyword in keywords)


class RunLog:
    """Collect row-level decisions so accounting reviewers can audit the run."""

    def __init__(self):
        self.records = []

    def add(self, level, category, message, row_number="", raw_row=""):
        self.records.append(
            {
                "Level": level,
                "Category": category,
                "Message": message,
                "Source Row": row_number,
                "Raw Row": raw_row,
            }
        )

    def to_frame(self):
        return pd.DataFrame(
            self.records,
            columns=["Level", "Category", "Message", "Source Row", "Raw Row"],
        )
